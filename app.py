import streamlit as st
import requests
from datetime import datetime, date
from zoneinfo import ZoneInfo
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =========================
# Supabase設定
# =========================
SUPABASE_URL = st.secrets.get("SUPABASE_URL", "https://uguwadlazrawhxkdtkhl.supabase.co")
SUPABASE_KEY = st.secrets.get("SUPABASE_KEY", "sb_publishable_r6wvPjR4FWfoJAWEKTiB5A_A-w1UWZv")

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

APP_PASSWORD = st.secrets.get("APP_PASSWORD", "mahjong")
ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "19831219")


# =========================
# Supabase API
# =========================
def api_get(table, params=None):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    response = requests.get(url, headers=HEADERS, params=params or {})
    if response.status_code >= 400:
        st.error(response.text)
        return []
    return response.json()


def api_post(table, data):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    response = requests.post(url, headers=HEADERS, json=data)
    if response.status_code >= 400:
        st.error(response.text)
        return None
    return response.json()


def api_upsert(table, data, on_conflict):
    """Supabaseへupsertする。設定値やチップ枚数の保存に使う。"""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = dict(HEADERS)
    headers["Prefer"] = "resolution=merge-duplicates,return=representation"
    response = requests.post(
        url,
        headers=headers,
        params={"on_conflict": on_conflict},
        json=data,
    )
    if response.status_code >= 400:
        st.error(response.text)
        return None
    return response.json()


def api_patch(table, row_id, data):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    params = {"id": f"eq.{row_id}"}
    response = requests.patch(url, headers=HEADERS, params=params, json=data)
    if response.status_code >= 400:
        st.error(response.text)
        return None
    return response.json()


def api_delete(table, row_id):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    params = {"id": f"eq.{row_id}"}
    response = requests.delete(url, headers=HEADERS, params=params)
    if response.status_code >= 400:
        st.error(response.text)
        return False
    return True


def api_delete_where(table, params):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    response = requests.delete(url, headers=HEADERS, params=params)
    if response.status_code >= 400:
        st.error(response.text)
        return False
    return True


# =========================
# アプリ設定・点数計算保存
# =========================
def get_app_setting(setting_key, default_value="0"):
    rows = api_get(
        "app_settings",
        {
            "select": "setting_key,setting_value",
            "setting_key": f"eq.{setting_key}",
            "limit": "1",
        },
    )
    if not rows:
        return default_value
    return rows[0].get("setting_value", default_value)


def get_app_setting_int(setting_key, default_value=0):
    try:
        return int(get_app_setting(setting_key, str(default_value)))
    except Exception:
        return int(default_value)


def normalize_rate_value(value, step):
    """過去に保存された 1 などの不正な初期値を0に戻す。"""
    try:
        value = int(value)
    except Exception:
        return 0
    if value < 0:
        return 0
    if step > 0 and value % step != 0:
        return 0
    return value


def save_app_setting(setting_key, setting_value):
    result = api_upsert(
        "app_settings",
        {
            "setting_key": str(setting_key),
            "setting_value": str(setting_value),
            "updated_at": datetime.now(ZoneInfo("Asia/Tokyo")).isoformat(),
        },
        "setting_key",
    )
    return result is not None


def point_calc_scope_key(session_id):
    return f"session_{session_id}" if session_id is not None else "all"


def get_saved_chip_counts(scope_key):
    rows = api_get(
        "point_calc_chips",
        {
            "select": "scope_key,player_id,chip_count",
            "scope_key": f"eq.{scope_key}",
        },
    )
    saved = {}
    for row in rows:
        try:
            saved[int(row["player_id"])] = int(row.get("chip_count") or 0)
        except Exception:
            pass
    return saved


def save_chip_count(scope_key, player_id, chip_count):
    result = api_upsert(
        "point_calc_chips",
        {
            "scope_key": str(scope_key),
            "player_id": int(player_id),
            "chip_count": int(chip_count),
            "updated_at": datetime.now(ZoneInfo("Asia/Tokyo")).isoformat(),
        },
        "scope_key,player_id",
    )
    return result is not None


# =========================
# データ操作：プレイヤー
# =========================
def get_players(include_hidden=False):
    select_cols = "id,name,created_at,is_active"
    params = {"select": select_cols, "order": "id.asc"}
    if not include_hidden:
        params["is_active"] = "eq.true"
    return api_get("players", params)


def add_player(name):
    name = name.strip()
    if not name:
        return False, "名前を入力してください。"

    players = get_players(include_hidden=True)
    active_same = [p for p in players if p.get("name") == name and p.get("is_active", True)]
    if active_same:
        return False, "同じ名前がすでに登録されています。"

    result = api_post("players", {"name": name, "is_active": True})
    if result is None:
        return False, "登録に失敗しました。"
    return True, f"{name} を登録しました。"


def update_player_name(player_id, new_name):
    new_name = new_name.strip()
    if not new_name:
        return False, "名前を入力してください。"
    result = api_patch("players", player_id, {"name": new_name})
    if result is None:
        return False, "名前変更に失敗しました。"
    return True, "名前を変更しました。"


def set_player_active(player_id, is_active):
    result = api_patch("players", player_id, {"is_active": is_active})
    if result is None:
        return False, "更新に失敗しました。"
    return True, "更新しました。"


def player_has_results(player_id):
    rows = api_get("game_results", {"select": "id", "player_id": f"eq.{player_id}", "limit": "1"})
    return len(rows) > 0


def delete_player(player_id):
    if player_has_results(player_id):
        return False, "このメンバーは対戦データがあるため削除できません。過去データを残すため、非表示を使ってください。"
    ok = api_delete("players", player_id)
    if not ok:
        return False, "削除に失敗しました。"
    return True, "削除しました。"


# =========================
# データ操作：対戦会（セッション）
# =========================
def get_sessions(include_finished=True):
    params = {"select": "id,title,session_date,status,created_at,finalized_at", "order": "session_date.desc,id.desc"}
    if not include_finished:
        params["status"] = "eq.active"
    return api_get("match_sessions", params)


def get_session(session_id):
    rows = api_get("match_sessions", {"select": "id,title,session_date,status,created_at,finalized_at", "id": f"eq.{session_id}", "limit": "1"})
    return rows[0] if rows else None


def create_session(title, session_date, player_ids):
    title = title.strip() or f"{session_date} 麻雀"
    created = api_post("match_sessions", {
        "title": title,
        "session_date": str(session_date),
        "status": "active",
    })
    if not created:
        return None
    session_id = created[0]["id"]
    rows = [{"session_id": session_id, "player_id": int(pid)} for pid in player_ids]
    if rows:
        api_post("session_players", rows)
    return session_id


def update_session_status(session_id, status):
    data = {"status": status}
    if status == "finished":
        data["finalized_at"] = datetime.now(ZoneInfo("Asia/Tokyo")).isoformat()
    return api_patch("match_sessions", session_id, data)


def get_session_players(session_id):
    rows = api_get(
        "session_players",
        {
            "select": "player_id,players(id,name,is_active)",
            "session_id": f"eq.{session_id}",
            "order": "player_id.asc",
        },
    )
    players = []
    for r in rows:
        p = r.get("players") or {}
        if p:
            players.append({"id": p.get("id"), "name": p.get("name"), "is_active": p.get("is_active", True)})
    return players


def add_player_to_session(session_id, player_id):
    current_players = get_session_players(session_id)
    current_ids = {int(p["id"]) for p in current_players if p.get("id") is not None}
    if int(player_id) in current_ids:
        return False, "このメンバーはすでに現在の対戦会に参加しています。"

    result = api_post("session_players", {
        "session_id": int(session_id),
        "player_id": int(player_id),
    })
    if result is None:
        return False, "対戦会への追加に失敗しました。"
    return True, "現在の対戦会に追加しました。"


def resume_session(session_id):
    result = api_patch("match_sessions", session_id, {
        "status": "active",
        "finalized_at": None,
    })
    if result is None:
        return False, "対戦会の再開に失敗しました。"
    return True, "対戦会を再開しました。"


def get_active_current_session():
    session_id = st.session_state.get("current_session_id")
    if not session_id:
        return None
    session = get_session(session_id)
    if not session or session.get("status") != "active":
        return None
    return session


def get_session_label(session):
    if not session:
        return "未設定"
    return f"{session.get('session_date') or ''}　{session.get('title') or '対戦会'}"


def get_session_member_ids(session_id):
    return {int(p["id"]) for p in get_session_players(session_id) if p.get("id") is not None}


def get_session_game_count(session_id):
    rows = api_get("games", {"select": "id", "session_id": f"eq.{session_id}"})
    return len(rows)


def get_next_game_no(session_id=None):
    params = {"select": "game_no", "order": "game_no.desc", "limit": "1"}
    if session_id is not None:
        params["session_id"] = f"eq.{session_id}"
    games = api_get("games", params)
    if not games:
        return 1
    return int(games[0]["game_no"]) + 1


def save_game(points, memo, session_id=None):
    game_no = get_next_game_no(session_id)
    game_data = {"game_no": game_no, "memo": memo}
    if session_id is not None:
        game_data["session_id"] = int(session_id)
    game = api_post("games", game_data)
    if not game:
        return False

    game_id = game[0]["id"]
    rows = []
    for player_id, point in points.items():
        rows.append({"game_id": game_id, "player_id": player_id, "point": int(point)})

    result = api_post("game_results", rows)
    return result is not None


def get_results(session_id=None):
    if session_id is None:
        select = "id,point,created_at,games(id,game_no,memo,created_at,session_id,match_sessions(id,title,session_date,status)),players(id,name)"
        params = {"select": select, "order": "id.asc"}
    else:
        select = "id,point,created_at,games!inner(id,game_no,memo,created_at,session_id,match_sessions(id,title,session_date,status)),players(id,name)"
        params = {"select": select, "games.session_id": f"eq.{session_id}", "order": "id.asc"}

    rows = api_get("game_results", params)

    results = []
    for row in rows:
        game = row.get("games") or {}
        session = game.get("match_sessions") or {}
        player = row.get("players") or {}
        results.append({
            "result_id": row.get("id"),
            "game_id": game.get("id"),
            "game_no": game.get("game_no"),
            "memo": game.get("memo"),
            "played_at": game.get("created_at"),
            "session_id": game.get("session_id"),
            "session_title": session.get("title") if session else None,
            "session_date": session.get("session_date") if session else None,
            "session_status": session.get("status") if session else None,
            "player_id": player.get("id"),
            "name": player.get("name"),
            "point": row.get("point", 0),
        })
    return results


def clear_score_data():
    # 名前マスタ(players)は消さず、対戦会・対戦結果だけ削除する
    if not api_delete_where("game_results", {"id": "gte.0"}):
        return False, "点数データの削除に失敗しました。"
    if not api_delete_where("games", {"id": "gte.0"}):
        return False, "対戦履歴の削除に失敗しました。"
    api_delete_where("session_players", {"session_id": "gte.0"})
    api_delete_where("match_sessions", {"id": "gte.0"})
    return True, "点数一覧・ランキング・個人成績・対戦履歴・対戦会データを全て削除しました。"


def delete_single_game(game_id):
    if not api_delete_where("game_results", {"game_id": f"eq.{game_id}"}):
        return False, "この対戦の点数データ削除に失敗しました。"
    if not api_delete("games", game_id):
        return False, "この対戦履歴の削除に失敗しました。"
    return True, "指定した対戦を削除しました。"


def get_game_count(session_id=None):
    params = {"select": "id"}
    if session_id is not None:
        params["session_id"] = f"eq.{session_id}"
    rows = api_get("games", params)
    return len(rows)


def get_result_count():
    rows = api_get("game_results", {"select": "id"})
    return len(rows)


# =========================
# 集計
# =========================
def game_label_from_row(row):
    no = row.get("game_no")
    session_date = row.get("session_date")
    if session_date:
        try:
            d = datetime.strptime(session_date, "%Y-%m-%d")
            return f"{d.month}/{d.day} {no}回戦"
        except Exception:
            return f"{session_date} {no}回戦"
    return f"{no}回戦"


def make_score_table(results):
    if not results:
        return []
    names = []
    for row in results:
        if row["name"] not in names:
            names.append(row["name"])

    labels = []
    label_order = []
    for row in results:
        label = game_label_from_row(row)
        if label not in labels:
            labels.append(label)
            label_order.append((row.get("session_date") or "", row.get("game_no") or 0, label))
    labels = [x[2] for x in sorted(label_order, key=lambda x: (x[0], x[1]))]

    table = []
    for name in names:
        line = {"名前": name}
        total = 0
        for label in labels:
            value = ""
            for row in results:
                if row["name"] == name and game_label_from_row(row) == label:
                    value = row["point"]
                    total += int(row["point"])
                    break
            line[label] = value
        line["累計"] = total
        table.append(line)
    table.sort(key=lambda x: x["累計"], reverse=True)
    return table


def make_point_calculation_base(results, session_id=None):
    """点数計算用の元データ。対戦会指定時は、その対戦会の全参加者を0点でも表示する。"""
    totals = {}
    names = {}

    for row in results:
        pid = row.get("player_id")
        if pid is None:
            continue
        pid = int(pid)
        names[pid] = row.get("name") or ""
        totals[pid] = totals.get(pid, 0) + int(row.get("point") or 0)

    # 対戦会別の場合は、チップ入力のためにその対戦会の参加者全員を出す
    if session_id is not None:
        for player in get_session_players(session_id):
            pid = player.get("id")
            if pid is None:
                continue
            pid = int(pid)
            names[pid] = player.get("name") or names.get(pid, "")
            totals.setdefault(pid, 0)

    rows = []
    for pid, name in names.items():
        rows.append({
            "player_id": pid,
            "名前": name,
            "累計": int(totals.get(pid, 0)),
        })

    rows.sort(key=lambda x: x["累計"], reverse=True)
    return rows


def group_results_by_game(results):
    games = {}
    for row in results:
        game_id = row.get("game_id")
        if game_id is None:
            continue
        if game_id not in games:
            games[game_id] = {
                "game_id": game_id,
                "game_no": row.get("game_no"),
                "memo": row.get("memo") or "",
                "played_at": row.get("played_at") or "",
                "session_id": row.get("session_id"),
                "session_title": row.get("session_title"),
                "session_date": row.get("session_date"),
                "members": [],
            }
        games[game_id]["members"].append({
            "player_id": row.get("player_id"),
            "name": row.get("name") or "",
            "point": int(row.get("point") or 0),
        })
    return games


def add_rank_to_games(games):
    for game in games.values():
        members = sorted(game["members"], key=lambda x: x["point"], reverse=True)
        prev_point = None
        current_rank = 0
        for idx, member in enumerate(members, start=1):
            if prev_point is None or member["point"] != prev_point:
                current_rank = idx
            member["rank"] = current_rank
            prev_point = member["point"]
        game["members"] = members
    return games


def build_player_stats(results):
    games = add_rank_to_games(group_results_by_game(results))
    stats = {}
    for game in games.values():
        member_count = len(game["members"])
        for member in game["members"]:
            pid = member["player_id"]
            name = member["name"]
            point = int(member["point"])
            rank = int(member.get("rank") or 0)
            if pid not in stats:
                stats[pid] = {
                    "player_id": pid,
                    "名前": name,
                    "対戦数": 0,
                    "トップ回数": 0,
                    "2着回数": 0,
                    "3着回数": 0,
                    "ラス回数": 0,
                    "プラス回数": 0,
                    "マイナス回数": 0,
                    "順位合計": 0,
                    "合計点": 0,
                    "最高点": point,
                    "最低点": point,
                }
            s = stats[pid]
            s["名前"] = name
            s["対戦数"] += 1
            s["合計点"] += point
            s["順位合計"] += rank
            s["最高点"] = max(s["最高点"], point)
            s["最低点"] = min(s["最低点"], point)
            if point > 0:
                s["プラス回数"] += 1
            elif point < 0:
                s["マイナス回数"] += 1
            if rank == 1:
                s["トップ回数"] += 1
            elif rank == 2:
                s["2着回数"] += 1
            elif rank == 3:
                s["3着回数"] += 1
            if rank == member_count:
                s["ラス回数"] += 1
    return stats, games


def make_enhanced_ranking(results):
    stats, _ = build_player_stats(results)
    table = []
    for data in stats.values():
        count = data["対戦数"]
        avg = data["合計点"] / count if count else 0
        avg_rank = data["順位合計"] / count if count else 0
        plus_rate = data["プラス回数"] / count * 100 if count else 0
        top_rate = data["トップ回数"] / count * 100 if count else 0
        last_rate = data["ラス回数"] / count * 100 if count else 0
        table.append({
            "順位": 0,
            "名前": data["名前"],
            "対戦数": count,
            "合計点": data["合計点"],
            "平均点": round(avg, 1),
            "平均順位": round(avg_rank, 2),
            "プラス率": f"{plus_rate:.1f}%",
            "トップ回数": data["トップ回数"],
            "トップ率": f"{top_rate:.1f}%",
            "ラス回数": data["ラス回数"],
            "ラス率": f"{last_rate:.1f}%",
            "最高点": data["最高点"],
            "最低点": data["最低点"],
        })
    table.sort(key=lambda x: (x["合計点"], x["平均点"]), reverse=True)
    for i, row in enumerate(table, start=1):
        row["順位"] = i
    return table


def make_personal_detail(results, target_name):
    stats, games = build_player_stats(results)
    personal_rows = []
    cumulative = 0
    for game in sorted(games.values(), key=lambda g: ((g.get("session_date") or ""), g.get("game_no") or 0)):
        members = game["members"]
        target = next((m for m in members if m["name"] == target_name), None)
        if not target:
            continue
        cumulative += int(target["point"])
        label = game_label_from_row({"game_no": game.get("game_no"), "session_date": game.get("session_date")})
        personal_rows.append({
            "回戦": label,
            "順位": target.get("rank"),
            "点数": int(target["point"]),
            "累計": cumulative,
            "メモ": game.get("memo") or "",
        })
    if not personal_rows:
        return None, [], []
    player_stat = None
    for s in stats.values():
        if s["名前"] == target_name:
            player_stat = s
            break
    chart_data = [{"回戦": r["回戦"], "累計": r["累計"]} for r in personal_rows]
    return player_stat, personal_rows, chart_data


def make_matchup_table(results, target_name):
    _, games = build_player_stats(results)
    matchup = {}
    for game in games.values():
        members = game["members"]
        me = next((m for m in members if m["name"] == target_name), None)
        if not me:
            continue
        for opp in members:
            if opp["name"] == target_name:
                continue
            name = opp["name"]
            if name not in matchup:
                matchup[name] = {
                    "相手": name,
                    "同卓回数": 0,
                    "自分合計": 0,
                    "相手合計": 0,
                    "勝ち回数": 0,
                    "負け回数": 0,
                    "自分順位合計": 0,
                    "相手順位合計": 0,
                }
            m = matchup[name]
            m["同卓回数"] += 1
            m["自分合計"] += int(me["point"])
            m["相手合計"] += int(opp["point"])
            m["自分順位合計"] += int(me.get("rank") or 0)
            m["相手順位合計"] += int(opp.get("rank") or 0)
            if int(me["point"]) > int(opp["point"]):
                m["勝ち回数"] += 1
            elif int(me["point"]) < int(opp["point"]):
                m["負け回数"] += 1

    table = []
    for m in matchup.values():
        count = m["同卓回数"]
        win_rate = m["勝ち回数"] / count * 100 if count else 0
        table.append({
            "相手": m["相手"],
            "同卓回数": count,
            "自分合計": m["自分合計"],
            "相手合計": m["相手合計"],
            "差分": m["自分合計"] - m["相手合計"],
            "自分平均": round(m["自分合計"] / count, 1) if count else 0,
            "相手平均": round(m["相手合計"] / count, 1) if count else 0,
            "勝ち回数": m["勝ち回数"],
            "勝率": f"{win_rate:.1f}%",
            "自分平均順位": round(m["自分順位合計"] / count, 2) if count else 0,
            "相手平均順位": round(m["相手順位合計"] / count, 2) if count else 0,
        })
    table.sort(key=lambda x: (x["差分"], x["同卓回数"]), reverse=True)
    return table


def make_history_table(results):
    games = add_rank_to_games(group_results_by_game(results))
    history = []
    for game in sorted(games.values(), key=lambda g: ((g.get("session_date") or ""), g.get("game_no") or 0), reverse=True):
        session_name = game.get("session_title") or "未分類"
        session_date = game.get("session_date") or ""
        for m in game["members"]:
            history.append({
                "対戦会": session_name,
                "日付": session_date,
                "回戦": game.get("game_no"),
                "順位": m.get("rank"),
                "名前": m.get("name"),
                "点数": m.get("point"),
                "メモ": game.get("memo") or "",
            })
    return history


def make_game_summary(results):
    if not results:
        return []
    games = add_rank_to_games(group_results_by_game(results))
    summary = []
    for game in games.values():
        members_sorted = sorted(game["members"], key=lambda x: x["point"], reverse=True)
        member_text = " / ".join([f"{m['name']} {m['point']:+d}" for m in members_sorted])
        session_date = game.get("session_date") or ""
        title = game.get("session_title") or "未分類"
        label = f"{session_date} {title}　{game['game_no']}回戦　{member_text}"
        if game["memo"]:
            label += f"　メモ：{game['memo']}"
        summary.append({
            "game_id": game["game_id"],
            "game_no": game["game_no"],
            "session_id": game.get("session_id"),
            "label": label,
            "members": members_sorted,
            "memo": game["memo"],
            "played_at": game["played_at"],
        })
    summary.sort(key=lambda x: (x.get("played_at") or "", x.get("game_no") or 0), reverse=True)
    return summary


def build_dashboard_metrics(results, players):
    ranking = make_enhanced_ranking(results)
    games = group_results_by_game(results)
    active_count = len(players)
    recent_game = make_game_summary(results)
    recent_label = recent_game[0]["label"] if recent_game else "まだ対戦がありません"
    return {
        "総対戦数": len(games),
        "登録メンバー": active_count,
        "ランキング": ranking,
        "直近対戦": recent_label,
    }


def make_excel_file(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "点数一覧"

    sheets = {
        "点数一覧": make_score_table(results),
        "ランキング": make_enhanced_ranking(results),
        "対戦履歴": make_history_table(results),
        "個人成績": make_enhanced_ranking(results),
    }

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        if not rows:
            ws.append(["データなし"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 28)
        ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


# =========================
# UI部品
# =========================
def zero_fill_point(current_player_id, selected_players):
    other_sum = 0
    for other in selected_players:
        if other["id"] == current_player_id:
            continue
        other_sum += int(st.session_state.get(f"manual_point_{other['id']}", 0))
    st.session_state[f"manual_point_{current_player_id}"] = -other_sum


def menu_button(label, icon, key):
    return st.button(f"{icon}\n\n{label}", key=key, use_container_width=True)


def back_button():
    if st.button("← メニューへ戻る", key="back_home"):
        go("home")


def select_result_scope(results_required=True):
    sessions = get_sessions(include_finished=True)
    options = ["全期間"]
    labels_to_ids = {"全期間": None}
    default_index = 0
    default_session_id = st.session_state.get("result_scope_default_session_id")

    for s in sessions:
        label = f"{s.get('session_date') or ''}　{s.get('title') or '対戦会'}"
        if s.get("status") == "active":
            label += "（進行中）"
        options.append(label)
        labels_to_ids[label] = s["id"]
        if default_session_id is not None and int(s["id"]) == int(default_session_id):
            default_index = len(options) - 1

    selected = st.selectbox("表示範囲", options, index=default_index)
    return labels_to_ids[selected]


def render_ranking_cards(ranking):
    if not ranking:
        st.info("まだランキングデータがありません。")
        return
    top = ranking[0]
    with st.container(border=True):
        st.markdown("### 🏆 現在1位")
        c1, c2 = st.columns([2, 1])
        c1.markdown(f"## {top['名前']}")
        c2.metric("合計点", f"{top['合計点']:+d}")
    others = ranking[1:4]
    if others:
        st.markdown("#### 2位〜4位")
        for row in others:
            with st.container(border=True):
                c1, c2, c3 = st.columns([0.7, 2.0, 1.1])
                c1.markdown(f"**{row['順位']}位**")
                c2.markdown(f"**{row['名前']}**")
                c3.markdown(f"**{row['合計点']:+d}**")


# =========================
# 画面設定・CSS
# =========================
st.set_page_config(page_title="麻雀スコア管理", page_icon="🀄", layout="wide")

st.markdown(
    """
    <style>
    html, body, [class*="css"] { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; }
    .block-container { padding-top: 1.1rem; padding-bottom: 2rem; max-width: 760px; }
    h1 { font-size: 2.2rem !important; margin-bottom: .8rem !important; font-weight: 800 !important; }
    h2, h3 { margin-top: 1rem !important; font-weight: 800 !important; }
    .stButton > button { border-radius: 12px; border: 1px solid #d1d5db; background: #ffffff; color: #111827; font-weight: 700; transition: 0.15s; }
    .stButton > button:hover { border-color: #ff4b4b; color: #ff4b4b; }
    .stButton > button[kind="primary"] { background: #ff4b4b; color: white; border-color: #ff4b4b; min-height: 54px; font-size: 1.05rem; border-radius: 16px; }
    div[data-testid="column"] .stButton > button[kind="secondary"] { min-height: 84px; white-space: pre-line; }
    div[data-testid="stVerticalBlockBorderWrapper"] { border-radius: 16px !important; border-color: #e5e7eb !important; box-shadow: 0 1px 2px rgba(0,0,0,0.03); }
    .member-id-label { color: #6b7280; font-size: 0.82rem; line-height: 1.1; margin-bottom: 0.18rem; }
    .member-name-label { color: #111827; font-size: 1.25rem; font-weight: 800; line-height: 1.15; word-break: keep-all; overflow-wrap: anywhere; }
    .button-spacer { height: 1.25rem; }
    .inline-panel { margin-top: .5rem; margin-bottom: .2rem; color: #6b7280; font-size: .85rem; font-weight: 700; }
    div[data-testid="stVerticalBlockBorderWrapper"] .stButton > button { min-height: 40px !important; height: 40px !important; padding: 0 .25rem !important; font-size: .9rem !important; border-radius: 10px !important; white-space: nowrap !important; }
    div[data-testid="stVerticalBlockBorderWrapper"] input { min-height: 40px !important; height: 40px !important; border-radius: 10px !important; }
    .score-card { background: #ffffff; border: 1px solid #e5e7eb; border-radius: 14px; padding: 10px 12px; margin-bottom: 8px; }
    .score-order { color: #6b7280; font-size: .82rem; font-weight: 700; margin-bottom: 2px; }
    .score-name { font-size: 1.15rem; font-weight: 800; color: #111827; }
    @media (max-width: 640px) {
        .block-container { padding-left: .85rem; padding-right: .85rem; max-width: 100% !important; }
        h1 { font-size: 1.75rem !important; }
        h2, h3 { font-size: 1.35rem !important; }
        div[data-testid="column"] .stButton > button[kind="secondary"] { min-height: 64px; }
        .member-id-label { font-size: .72rem; }
        .member-name-label { font-size: 1.05rem; }
        .button-spacer { height: 1.05rem; }
        div[data-testid="stVerticalBlockBorderWrapper"] { padding: .15rem !important; }
        div[data-testid="stVerticalBlockBorderWrapper"] .stButton > button { min-height: 34px !important; height: 34px !important; padding: 0 .12rem !important; font-size: .72rem !important; border-radius: 8px !important; }
        div[data-testid="stVerticalBlockBorderWrapper"] input { min-height: 36px !important; height: 36px !important; font-size: .9rem !important; }
        .score-name { font-size: 1.05rem; }
        .score-order { font-size: .76rem; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# =========================
# 簡易ログイン
# =========================
if "login" not in st.session_state:
    st.session_state.login = False

if not st.session_state.login:
    st.title("🀄 麻雀スコア管理")
    st.write("合言葉を入力してください。")
    password = st.text_input("合言葉", type="password")
    if st.button("ログイン", type="primary"):
        if password == APP_PASSWORD:
            st.session_state.login = True
            st.rerun()
        else:
            st.error("合言葉が違います。")
    st.stop()


# =========================
# ページ状態
# =========================
if "page" not in st.session_state:
    st.session_state.page = "home"
if "delete_confirm_id" not in st.session_state:
    st.session_state.delete_confirm_id = None
if "edit_player_id" not in st.session_state:
    st.session_state.edit_player_id = None
if "clear_scores_step" not in st.session_state:
    st.session_state.clear_scores_step = 0
if "delete_game_confirm_id" not in st.session_state:
    st.session_state.delete_game_confirm_id = None
if "selected_player_ids" not in st.session_state:
    st.session_state.selected_player_ids = []
if "setup_session_player_ids" not in st.session_state:
    st.session_state.setup_session_player_ids = []
if "current_session_id" not in st.session_state:
    st.session_state.current_session_id = None
if "save_complete" not in st.session_state:
    st.session_state.save_complete = False
if "finish_confirm_session_id" not in st.session_state:
    st.session_state.finish_confirm_session_id = None
if "resume_confirm_session_id" not in st.session_state:
    st.session_state.resume_confirm_session_id = None
if "result_scope_default_session_id" not in st.session_state:
    st.session_state.result_scope_default_session_id = None
if "point_calc_score_rate" not in st.session_state:
    st.session_state.point_calc_score_rate = get_app_setting_int("point_calc_score_rate", 0)
if "point_calc_chip_rate" not in st.session_state:
    st.session_state.point_calc_chip_rate = get_app_setting_int("point_calc_chip_rate", 0)


def clear_hand_selection():
    st.session_state.selected_player_ids = []
    for key in list(st.session_state.keys()):
        if str(key).startswith("manual_point_"):
            del st.session_state[key]


def go(page):
    st.session_state.page = page
    st.rerun()


# =========================
# トップ画面
# =========================
if st.session_state.page == "home":
    st.title("🀄 麻雀スコア管理")
    st.caption("メニューを選択してください。")

    all_results = get_results()
    players = get_players()
    metrics = build_dashboard_metrics(all_results, players)

    c1, c2 = st.columns(2)
    c1.metric("総対戦数", f"{metrics['総対戦数']}戦")
    c2.metric("登録メンバー", f"{metrics['登録メンバー']}人")

    render_ranking_cards(metrics["ランキング"])

    with st.container(border=True):
        st.markdown("**直近の対戦**")
        st.write(metrics["直近対戦"])

    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        if menu_button("対戦会 設定/終了", "🗓️", "session_manage"):
            go("session_manage")
        if menu_button("対戦スタート", "🎮", "start"):
            clear_hand_selection()
            st.session_state.finish_confirm_session_id = None
            go("start")
        if menu_button("点数一覧", "📋", "score_list"):
            go("score_list")
        if menu_button("個人成績", "👤", "personal"):
            go("personal")
        if menu_button("過去の対戦履歴", "🕘", "history"):
            go("history")
    with col2:
        if menu_button("名前登録", "✏️", "players"):
            go("players")
        if menu_button("ランキング", "🏆", "ranking"):
            go("ranking")
        if menu_button("点数計算", "🧮", "point_calc"):
            go("point_calc")
        if menu_button("相性分析", "🤝", "matchup"):
            go("matchup")
        if menu_button("設定", "⚙️", "settings"):
            go("settings")


# =========================
# 対戦会 設定/終了
# =========================
elif st.session_state.page == "session_manage":
    st.title("🗓️ 対戦会 設定/終了")
    back_button()

    players = get_players()
    active_sessions = get_sessions(include_finished=False)
    all_sessions = get_sessions(include_finished=True)
    finished_sessions = [s for s in all_sessions if s.get("status") == "finished"]

    current_session = get_active_current_session()

    if current_session:
        st.subheader("現在の対戦会")
        session_id = current_session["id"]
        session_players = get_session_players(session_id)
        session_results = get_results(session_id=session_id)
        session_ranking = make_enhanced_ranking(session_results)

        with st.container(border=True):
            st.markdown(f"### {get_session_label(current_session)}")
            c1, c2, c3 = st.columns(3)
            c1.metric("参加者", f"{len(session_players)}人")
            c2.metric("対戦数", f"{get_session_game_count(session_id)}戦")
            c3.metric("状態", "進行中")

            if session_ranking:
                render_ranking_cards(session_ranking)
            else:
                st.info("まだ対戦結果は登録されていません。")

        # 進行中の対戦会がある間は、新しい対戦会は作れないようにする
        st.info("この対戦会が終了するまで、新しい対戦会は作成できません。")

        st.markdown("---")
        st.subheader("途中参加メンバーを追加")
        st.caption("対戦会の途中でも、まだ参加していない登録メンバーを追加できます。")

        joined_ids = get_session_member_ids(session_id)
        not_joined = [p for p in players if int(p["id"]) not in joined_ids]

        if not_joined:
            for p in not_joined:
                with st.container(border=True):
                    c1, c2 = st.columns([3, 1], gap="small")
                    with c1:
                        st.markdown(
                            f'<div class="member-id-label">ID: {p["id"]}</div><div class="member-name-label">{p["name"]}</div>',
                            unsafe_allow_html=True,
                        )
                    with c2:
                        st.markdown('<div class="button-spacer"></div>', unsafe_allow_html=True)
                        if st.button("参加追加", key=f"session_manage_join_{p['id']}", use_container_width=True):
                            ok, msg = add_player_to_session(session_id, p["id"])
                            if ok:
                                st.success(msg)
                                st.rerun()
                            else:
                                st.warning(msg)
        else:
            st.success("登録済みの表示メンバーは全員、この対戦会に参加しています。")

        st.markdown("---")
        st.subheader("対戦会の終了")
        st.caption("終了すると、この対戦会の累計結果画面へ移動します。後から再開することもできます。")

        if st.session_state.finish_confirm_session_id != session_id:
            if st.button("点数計算して終了", type="primary", use_container_width=True):
                st.session_state.finish_confirm_session_id = session_id
                st.rerun()
        else:
            st.warning("本当に終了してよろしいですか？")
            yes_col, no_col = st.columns(2, gap="small")
            with yes_col:
                if st.button("はい、終了します", key=f"finish_yes_manage_{session_id}", use_container_width=True):
                    update_session_status(session_id, "finished")
                    st.session_state.result_scope_default_session_id = session_id
                    st.session_state.current_session_id = None
                    st.session_state.finish_confirm_session_id = None
                    clear_hand_selection()
                    go("score_list")
            with no_col:
                if st.button("いいえ", key=f"finish_no_manage_{session_id}", use_container_width=True):
                    st.session_state.finish_confirm_session_id = None
                    st.rerun()

    else:
        st.info("現在進行中として選択されている対戦会はありません。")

        if active_sessions:
            st.subheader("進行中の対戦会を選択")
            st.caption("進行中の対戦会があるため、新しい対戦会は作成できません。既存の対戦会を選択してください。")
            for s in active_sessions:
                with st.container(border=True):
                    c1, c2 = st.columns([3, 1], gap="small")
                    with c1:
                        st.markdown(f"**{get_session_label(s)}**")
                        st.caption(f"現在 {get_session_game_count(s['id'])} 戦")
                    with c2:
                        if st.button("選択", key=f"select_active_session_{s['id']}", use_container_width=True):
                            st.session_state.current_session_id = s["id"]
                            clear_hand_selection()
                            st.session_state.finish_confirm_session_id = None
                            st.rerun()
        else:
            st.markdown("---")
            st.subheader("新しい対戦会を作成")

            # 日付を変更したら、対戦名の日付部分も自動で変更する
            if "new_session_date" not in st.session_state:
                st.session_state.new_session_date = date.today()
            if "new_session_title_auto" not in st.session_state:
                st.session_state.new_session_title_auto = True
            if "new_session_title" not in st.session_state:
                st.session_state.new_session_title = f"{st.session_state.new_session_date.strftime('%Y/%m/%d')} 麻雀"

            previous_session_date = st.session_state.new_session_date

            session_date = st.date_input(
                "日付",
                key="new_session_date",
            )

            auto_title = f"{session_date.strftime('%Y/%m/%d')} 麻雀"
            previous_auto_title = f"{previous_session_date.strftime('%Y/%m/%d')} 麻雀"

            # 対戦名が自動生成のまま、または前回の自動生成名のままなら、日付変更に合わせて更新
            if st.session_state.new_session_title_auto or st.session_state.new_session_title == previous_auto_title:
                st.session_state.new_session_title = auto_title
                st.session_state.new_session_title_auto = True

            title = st.text_input(
                "対戦名",
                key="new_session_title",
            )

            # 手入力で対戦名を変えた場合は、以後その手入力を優先
            st.session_state.new_session_title_auto = title == auto_title

            if len(players) < 4:
                st.warning("先に4人以上の名前を登録してください。")
            else:
                st.caption("最初に参加するメンバーを全員選んでください。途中参加はこの画面から追加できます。")
                valid_ids = {p["id"] for p in players}
                st.session_state.setup_session_player_ids = [pid for pid in st.session_state.setup_session_player_ids if pid in valid_ids]
                st.info(f"現在 {len(st.session_state.setup_session_player_ids)} 人 選択中")

                reset_col, _ = st.columns([1.3, 3.7])
                with reset_col:
                    if st.button("参加者リセット", use_container_width=True):
                        st.session_state.setup_session_player_ids = []
                        st.rerun()

                for p in players:
                    pid = p["id"]
                    is_selected = pid in st.session_state.setup_session_player_ids
                    with st.container(border=True):
                        c1, c2 = st.columns([3, 1], gap="small")
                        with c1:
                            st.markdown(f"**{p['name']}**")
                            st.caption(f"ID:{pid}")
                        with c2:
                            if is_selected:
                                if st.button("解除", key=f"setup_unselect_{pid}", use_container_width=True):
                                    st.session_state.setup_session_player_ids.remove(pid)
                                    st.rerun()
                            else:
                                if st.button("参加", key=f"setup_select_{pid}", use_container_width=True):
                                    st.session_state.setup_session_player_ids.append(pid)
                                    st.rerun()

                if st.button("このメンバーで対戦会を開始", type="primary", use_container_width=True):
                    if len(st.session_state.setup_session_player_ids) < 4:
                        st.warning("参加者は4人以上必要です。")
                    else:
                        sid = create_session(title, session_date, st.session_state.setup_session_player_ids)
                        if sid:
                            st.session_state.current_session_id = sid
                            st.session_state.setup_session_player_ids = []
                            clear_hand_selection()
                            st.session_state.finish_confirm_session_id = None
                            st.success("対戦会を開始しました。")
                            go("start")
                        else:
                            st.error("対戦会の作成に失敗しました。")

    if finished_sessions:
        st.markdown("---")
        st.subheader("終了済みの対戦会を再開")
        st.caption("一度終了した対戦会も、確認後に再開できます。")

        for s in finished_sessions[:20]:
            with st.container(border=True):
                c1, c2 = st.columns([3, 1], gap="small")
                with c1:
                    st.markdown(f"**{get_session_label(s)}**")
                    st.caption(f"対戦数：{get_session_game_count(s['id'])}戦")
                with c2:
                    if st.button("再開", key=f"resume_open_{s['id']}", use_container_width=True):
                        st.session_state.resume_confirm_session_id = s["id"]
                        st.rerun()

                if st.session_state.resume_confirm_session_id == s["id"]:
                    st.warning("この対戦会を再開しますか？")
                    yes_col, no_col = st.columns(2, gap="small")
                    with yes_col:
                        if st.button("はい、再開します", key=f"resume_yes_{s['id']}", use_container_width=True):
                            ok, msg = resume_session(s["id"])
                            if ok:
                                st.session_state.current_session_id = s["id"]
                                st.session_state.resume_confirm_session_id = None
                                clear_hand_selection()
                                st.success(msg)
                                go("start")
                            else:
                                st.warning(msg)
                    with no_col:
                        if st.button("いいえ", key=f"resume_no_{s['id']}", use_container_width=True):
                            st.session_state.resume_confirm_session_id = None
                            st.rerun()


# =========================
# 名前登録
# =========================
elif st.session_state.page == "players":
    st.title("✏️ 名前登録")
    back_button()

    st.markdown("### 参加者を登録")
    new_name = st.text_input("名前", placeholder="例：小野", label_visibility="collapsed")
    if st.button("＋ 登録する", type="primary", use_container_width=True):
        ok, msg = add_player(new_name)
        if ok:
            st.success(msg)
            st.rerun()
        else:
            st.warning(msg)

    st.markdown("---")
    st.subheader("登録済みメンバー")

    players = get_players(include_hidden=False)
    hidden_players_all = get_players(include_hidden=True)
    hidden_players = [p for p in hidden_players_all if not p.get("is_active", True)]

    if players:
        for p in players:
            with st.container(border=True):
                name_col, change_col, hide_col, delete_col = st.columns([2.25, 1.05, 0.95, 0.85], gap="small")

                with name_col:
                    st.markdown(
                        f"""
                        <div class="member-id-label">ID: {p["id"]}</div>
                        <div class="member-name-label">{p["name"]}</div>
                        """,
                        unsafe_allow_html=True,
                    )

                with change_col:
                    st.markdown('<div class="button-spacer"></div>', unsafe_allow_html=True)
                    if st.button("名称変更", key=f"open_edit_{p['id']}", use_container_width=True):
                        st.session_state.edit_player_id = p["id"]
                        st.session_state.delete_confirm_id = None
                        st.rerun()

                with hide_col:
                    st.markdown('<div class="button-spacer"></div>', unsafe_allow_html=True)
                    if st.button("非表示", key=f"hide_{p['id']}", use_container_width=True):
                        ok, msg = set_player_active(p["id"], False)
                        if ok:
                            st.success("非表示にしました。")
                            st.rerun()
                        else:
                            st.warning(msg)

                with delete_col:
                    st.markdown('<div class="button-spacer"></div>', unsafe_allow_html=True)
                    if st.button("削除", key=f"delete_{p['id']}", use_container_width=True):
                        st.session_state.delete_confirm_id = p["id"]
                        st.session_state.edit_player_id = None
                        st.rerun()

                if st.session_state.edit_player_id == p["id"]:
                    st.markdown('<div class="inline-panel">名前を変更</div>', unsafe_allow_html=True)
                    edit_col, save_col, cancel_col = st.columns([3.0, 0.9, 0.9], gap="small")
                    with edit_col:
                        edited_name = st.text_input(
                            "新しい名前",
                            value=p["name"],
                            key=f"edit_name_{p['id']}",
                            label_visibility="collapsed",
                        )
                    with save_col:
                        if st.button("保存", key=f"save_edit_{p['id']}", use_container_width=True):
                            ok, msg = update_player_name(p["id"], edited_name)
                            if ok:
                                st.session_state.edit_player_id = None
                                st.success(msg)
                                st.rerun()
                            else:
                                st.warning(msg)
                    with cancel_col:
                        if st.button("取消", key=f"cancel_edit_{p['id']}", use_container_width=True):
                            st.session_state.edit_player_id = None
                            st.rerun()

                if st.session_state.delete_confirm_id == p["id"]:
                    st.warning(f"{p['name']} さんを削除しますか？")
                    yes_col, no_col, blank_col = st.columns([0.9, 0.9, 3.2], gap="small")
                    with yes_col:
                        if st.button("はい", key=f"delete_yes_{p['id']}", use_container_width=True):
                            ok, msg = delete_player(p["id"])
                            st.session_state.delete_confirm_id = None
                            if ok:
                                st.success(msg)
                                st.rerun()
                            else:
                                st.warning(msg)
                    with no_col:
                        if st.button("いいえ", key=f"delete_no_{p['id']}", use_container_width=True):
                            st.session_state.delete_confirm_id = None
                            st.rerun()
    else:
        st.info("表示中のメンバーはいません。")

    active_session_for_join = get_active_current_session()
    if active_session_for_join and players:
        st.markdown("---")
        st.subheader("現在の対戦会へ途中参加")
        st.caption("対戦会の途中でも、まだ参加していないメンバーを追加できます。")
        joined_ids = get_session_member_ids(active_session_for_join["id"])
        not_joined = [p for p in players if int(p["id"]) not in joined_ids]

        if not_joined:
            st.info(f"現在の対戦会：{get_session_label(active_session_for_join)}")
            for p in not_joined:
                with st.container(border=True):
                    c1, c2 = st.columns([3, 1], gap="small")
                    with c1:
                        st.markdown(f'<div class="member-id-label">ID: {p["id"]}</div><div class="member-name-label">{p["name"]}</div>', unsafe_allow_html=True)
                    with c2:
                        st.markdown('<div class="button-spacer"></div>', unsafe_allow_html=True)
                        if st.button("参加追加", key=f"join_current_session_{p['id']}", use_container_width=True):
                            ok, msg = add_player_to_session(active_session_for_join["id"], p["id"])
                            if ok:
                                st.success(msg)
                                st.rerun()
                            else:
                                st.warning(msg)
        else:
            st.success("登録済みの表示メンバーは全員、現在の対戦会に参加しています。")

    st.markdown("---")
    if hidden_players:
        with st.expander(f"非表示メンバー（{len(hidden_players)}人）"):
            for p in hidden_players:
                with st.container(border=True):
                    c1, c2 = st.columns([3.2, 0.8], gap="small")
                    with c1:
                        st.markdown(
                            f'<div class="member-id-label">ID: {p["id"]}</div><div class="member-name-label">{p["name"]}</div>',
                            unsafe_allow_html=True,
                        )
                    with c2:
                        st.markdown('<div class="button-spacer"></div>', unsafe_allow_html=True)
                        if st.button("復活", key=f"restore_{p['id']}", use_container_width=True):
                            ok, msg = set_player_active(p["id"], True)
                            if ok:
                                st.success("復活しました。")
                                st.rerun()
                            else:
                                st.warning(msg)


# =========================
# 対戦スタート
# =========================
elif st.session_state.page == "start":
    st.title("🎮 対戦スタート")
    back_button()

    if st.session_state.get("save_complete", False):
        clear_hand_selection()
        st.success("正常に登録されました。")
        st.info("次の半荘を登録する場合は、下のボタンから4人を選択してください。")
        if st.button("次の対戦へ", type="primary", use_container_width=True):
            st.session_state.save_complete = False
            clear_hand_selection()
            st.rerun()
        if st.button("点数一覧を確認する", use_container_width=True):
            st.session_state.save_complete = False
            st.session_state.result_scope_default_session_id = st.session_state.current_session_id
            go("score_list")
        st.stop()

    session_id = st.session_state.current_session_id
    session = get_session(session_id) if session_id else None

    if not session or session.get("status") != "active":
        st.warning("先に『対戦会 設定/終了』から、対戦会を設定してください。")
        if st.button("対戦会 設定/終了へ", type="primary", use_container_width=True):
            go("session_manage")
        st.stop()

    session_players = get_session_players(session_id)
    session_results = get_results(session_id=session_id)
    session_ranking = make_enhanced_ranking(session_results)

    st.subheader(f"{session.get('session_date')}　{session.get('title')}")
    c1, c2, c3 = st.columns(3)
    c1.metric("参加者", f"{len(session_players)}人")
    c2.metric("対戦数", f"{get_session_game_count(session_id)}戦")
    c3.metric("状態", "進行中")

    if session_ranking:
        render_ranking_cards(session_ranking)

    st.info("この対戦会を終了する場合は、ホームの『対戦会 設定/終了』から終了してください。")

    st.markdown("---")
    st.subheader("半荘を登録")

    if len(session_players) < 4:
        st.warning("この対戦会の参加者が4人未満です。『名前登録』から途中参加者を追加してください。")
        if st.button("名前登録へ", use_container_width=True):
            go("players")
        st.stop()

    id_to_player = {p["id"]: p for p in session_players}
    valid_ids = {p["id"] for p in session_players}
    st.session_state.selected_player_ids = [pid for pid in st.session_state.selected_player_ids if pid in valid_ids]
    selected_ids = st.session_state.selected_player_ids
    st.info(f"現在 {len(selected_ids)} / 4人 選択中")
    if selected_ids:
        st.success("選択中：" + " / ".join([id_to_player[pid]["name"] for pid in selected_ids]))

    reset_col, _ = st.columns([1.2, 3.8])
    with reset_col:
        if st.button("選択リセット", use_container_width=True):
            clear_hand_selection()
            st.rerun()

    for p in session_players:
        pid = p["id"]
        is_selected = pid in st.session_state.selected_player_ids
        with st.container(border=True):
            name_col, btn_col = st.columns([3.2, 1.0], gap="small")
            with name_col:
                order_text = f"　{selected_ids.index(pid) + 1}人目" if is_selected else ""
                st.markdown(f'<div class="member-id-label">ID: {pid}{order_text}</div><div class="member-name-label">{p["name"]}</div>', unsafe_allow_html=True)
            with btn_col:
                st.markdown('<div class="button-spacer"></div>', unsafe_allow_html=True)
                if is_selected:
                    if st.button("解除", key=f"unselect_{pid}", use_container_width=True):
                        st.session_state.selected_player_ids.remove(pid)
                        key = f"manual_point_{pid}"
                        if key in st.session_state:
                            st.session_state[key] = 0
                        st.rerun()
                else:
                    disabled = len(st.session_state.selected_player_ids) >= 4
                    if st.button("選択", key=f"select_{pid}", use_container_width=True, disabled=disabled):
                        st.session_state.selected_player_ids.append(pid)
                        st.rerun()

    if len(st.session_state.selected_player_ids) != 4:
        st.info("4人選択すると、半荘結果を入力できます。")
    else:
        selected_players = [id_to_player[pid] for pid in st.session_state.selected_player_ids]
        st.markdown("---")
        st.subheader(f"{get_next_game_no(session_id)}回戦の結果を入力")
        st.caption("全員手入力できます。点数は1刻みです。最後に入力する人は『集計』ボタンで合計0になる値を自動入力できます。")

        for i, player in enumerate(selected_players, start=1):
            with st.container(border=True):
                name_col, score_col, auto_col = st.columns([2.0, 1.35, 0.9], gap="small")
                with name_col:
                    st.markdown(f'<div class="score-order">{i}人目</div><div class="score-name">{player["name"]}</div>', unsafe_allow_html=True)
                with score_col:
                    key = f"manual_point_{player['id']}"
                    if key not in st.session_state:
                        st.session_state[key] = 0
                    st.number_input("点数", value=int(st.session_state[key]), step=1, key=key, label_visibility="collapsed")
                with auto_col:
                    st.markdown('<div class="button-spacer"></div>', unsafe_allow_html=True)
                    st.button("集計", key=f"zero_fill_{player['id']}", use_container_width=True, on_click=zero_fill_point, args=(player["id"], selected_players))

        total = sum(int(st.session_state.get(f"manual_point_{p['id']}", 0)) for p in selected_players)
        if total == 0:
            st.success("合計は0です。登録できます。")
        else:
            st.error(f"合計が {total} です。±0になっていないため登録できません。")

        memo = st.text_input("メモ", placeholder="例：1回目、南場で逆転 など")
        st.subheader("登録前確認")
        preview = []
        final_points = {}
        for p in selected_players:
            point = int(st.session_state.get(f"manual_point_{p['id']}", 0))
            final_points[p["id"]] = point
            preview.append({"名前": p["name"], "点数": point})
        preview.sort(key=lambda x: x["点数"], reverse=True)
        st.table(preview)

        if st.button("この対戦を登録する", type="primary", use_container_width=True):
            if total != 0:
                st.error("合計が±0になっていないため登録できません。")
            else:
                ok = save_game(final_points, memo, session_id=session_id)
                if ok:
                    st.session_state.save_complete = True
                    st.rerun()
                else:
                    st.error("登録に失敗しました。")


# =========================
# 点数一覧
# =========================
elif st.session_state.page == "score_list":
    st.title("📋 点数一覧")
    back_button()

    selected_session_id = select_result_scope()
    results = get_results(session_id=selected_session_id)
    table = make_score_table(results)

    if table:
        st.subheader("点数一覧")
        st.table(table)

        st.markdown("---")
        with st.expander("⚠️ 管理者メニュー", expanded=False):
            st.markdown("### 指定の対戦削除")
            st.caption("間違えて登録した対戦だけを削除できます。名前登録データは削除されません。")

            game_summary = make_game_summary(results)
            if game_summary:
                options = {g["label"]: g for g in game_summary}
                selected_label = st.selectbox("削除したい対戦を選択", list(options.keys()))
                selected_game = options[selected_label]

                st.markdown("**選択中の対戦**")
                st.table([{"名前": m["name"], "点数": m["point"]} for m in selected_game["members"]])
                if selected_game.get("memo"):
                    st.write(f"メモ：{selected_game['memo']}")

                if st.session_state.delete_game_confirm_id != selected_game["game_id"]:
                    if st.button("この対戦を削除", use_container_width=True):
                        st.session_state.delete_game_confirm_id = selected_game["game_id"]
                        st.rerun()
                else:
                    st.warning("この対戦を削除しますが、よろしいですか？")
                    yes_col, no_col = st.columns(2, gap="small")
                    with yes_col:
                        if st.button("はい、削除します", use_container_width=True):
                            ok, msg = delete_single_game(selected_game["game_id"])
                            st.session_state.delete_game_confirm_id = None
                            if ok:
                                st.success(msg)
                                st.rerun()
                            else:
                                st.warning(msg)
                    with no_col:
                        if st.button("いいえ", use_container_width=True):
                            st.session_state.delete_game_confirm_id = None
                            st.rerun()
            else:
                st.info("削除できる対戦データがありません。")
    else:
        st.info("まだデータがありません。")


# =========================
# ランキング
# =========================
elif st.session_state.page == "ranking":
    st.title("🏆 ランキング")
    back_button()
    selected_session_id = select_result_scope()
    results = get_results(session_id=selected_session_id)
    table = make_enhanced_ranking(results)
    if table:
        render_ranking_cards(table)
        st.markdown("---")
        st.subheader("総合ランキング")
        st.table(table)
    else:
        st.info("まだデータがありません。")


# =========================
# 個人成績
# =========================
elif st.session_state.page == "personal":
    st.title("👤 個人成績")
    back_button()
    selected_session_id = select_result_scope()
    results = get_results(session_id=selected_session_id)
    names = sorted(set(r["name"] for r in results if r["name"]))
    if not names:
        st.info("まだデータがありません。")
    else:
        target = st.selectbox("名前を選択", names)
        stat, personal_rows, chart_data = make_personal_detail(results, target)
        if not stat:
            st.info("データがありません。")
        else:
            count = stat["対戦数"]
            total = stat["合計点"]
            avg = total / count if count else 0
            avg_rank = stat["順位合計"] / count if count else 0
            top_rate = stat["トップ回数"] / count * 100 if count else 0
            last_rate = stat["ラス回数"] / count * 100 if count else 0
            plus_rate = stat["プラス回数"] / count * 100 if count else 0

            c1, c2, c3 = st.columns(3)
            c1.metric("対戦数", count)
            c2.metric("合計点", f"{total:+d}")
            c3.metric("平均点", f"{avg:.1f}")
            c4, c5, c6 = st.columns(3)
            c4.metric("平均順位", f"{avg_rank:.2f}")
            c5.metric("トップ率", f"{top_rate:.1f}%")
            c6.metric("ラス率", f"{last_rate:.1f}%")
            c7, c8, c9 = st.columns(3)
            c7.metric("プラス率", f"{plus_rate:.1f}%")
            c8.metric("最高点", f"{stat['最高点']:+d}")
            c9.metric("最低点", f"{stat['最低点']:+d}")

            st.markdown("---")
            st.subheader("順位内訳")
            st.table([{"トップ": stat["トップ回数"], "2着": stat["2着回数"], "3着": stat["3着回数"], "ラス": stat["ラス回数"]}])
            st.subheader("累計推移")
            st.line_chart({"累計": [r["累計"] for r in personal_rows]})
            st.subheader("直近10戦")
            st.table(list(reversed(personal_rows[-10:])))



# =========================
# 点数計算
# =========================
elif st.session_state.page == "point_calc":
    st.title("🧮 点数計算")
    back_button()

    st.caption("対戦会ごとの累計点に、点数ptとチップptを掛けて総合計を計算します。")

    selected_session_id = select_result_scope(results_required=False)
    results = get_results(session_id=selected_session_id)
    base_rows = make_point_calculation_base(results, session_id=selected_session_id)

    if not base_rows:
        st.info("選択した範囲には、まだ計算対象のメンバーがいません。")
    else:
        # Supabaseに保存されている最後の設定値を、画面初回表示時に読み込む。
        # ただし、過去に保存された「1」など刻み幅に合わない値は0に戻す。
        if "point_calc_score_rate" not in st.session_state:
            saved_score_rate = get_app_setting_int("point_calc_score_rate", 0)
            st.session_state.point_calc_score_rate = normalize_rate_value(saved_score_rate, 10)
            if saved_score_rate != st.session_state.point_calc_score_rate:
                save_app_setting("point_calc_score_rate", st.session_state.point_calc_score_rate)

        if "point_calc_chip_rate" not in st.session_state:
            saved_chip_rate = get_app_setting_int("point_calc_chip_rate", 0)
            st.session_state.point_calc_chip_rate = normalize_rate_value(saved_chip_rate, 100)
            if saved_chip_rate != st.session_state.point_calc_chip_rate:
                save_app_setting("point_calc_chip_rate", st.session_state.point_calc_chip_rate)

        st.subheader("ポイント設定")
        rate_col1, rate_col2 = st.columns(2, gap="small")
        with rate_col1:
            score_rate = st.number_input(
                "1点あたりのpt",
                min_value=0,
                value=int(st.session_state.get("point_calc_score_rate", 0)),
                step=10,
                key="point_calc_score_rate",
            )
        with rate_col2:
            chip_rate = st.number_input(
                "チップ1枚あたりのpt",
                min_value=0,
                value=int(st.session_state.get("point_calc_chip_rate", 0)),
                step=100,
                key="point_calc_chip_rate",
            )

        # 誰かが最後に変更した値を保存。次回以降、全員の画面で同じ値を使う。
        save_app_setting("point_calc_score_rate", int(score_rate))
        save_app_setting("point_calc_chip_rate", int(chip_rate))
        st.caption(f"現在の保存値：1点={int(score_rate)}pt / チップ1枚={int(chip_rate)}pt")

        st.markdown("---")
        st.subheader("計算表")
        st.caption("チップ枚数は選択した対戦会の全員分を入力してください。チップ合計が0にならない場合はエラー表示します。")

        scope_key = point_calc_scope_key(selected_session_id)
        saved_chips = get_saved_chip_counts(scope_key)

        header_cols = st.columns([1.55, 0.82, 1.05, 0.82, 1.05, 1.15], gap="small")
        headers = ["名前", "累計", f"点数pt\n(×{score_rate})", "チップ", f"チップpt\n(×{chip_rate})", "総合計"]
        for col, label in zip(header_cols, headers):
            with col:
                st.markdown(f"**{label}**")

        final_rows = []
        chip_total = 0

        for row in base_rows:
            pid = int(row["player_id"])
            name = row["名前"]
            total_score = int(row["累計"])
            score_points = total_score * int(score_rate)
            chip_key = f"chip_count_{scope_key}_{pid}"
            if chip_key not in st.session_state:
                st.session_state[chip_key] = int(saved_chips.get(pid, 0))

            c1, c2, c3, c4, c5, c6 = st.columns([1.55, 0.82, 1.05, 0.82, 1.05, 1.15], gap="small")
            with c1:
                st.write(name)
            with c2:
                st.write(f"{total_score:+d}")
            with c3:
                st.write(f"{score_points:+d}")
            with c4:
                chip_count = st.number_input(
                    "チップ",
                    value=int(st.session_state[chip_key]),
                    step=1,
                    key=chip_key,
                    label_visibility="collapsed",
                )

            chip_count = int(chip_count)
            # 入力されたチップ枚数も対戦会ごとに保存
            if int(saved_chips.get(pid, 0)) != chip_count:
                save_chip_count(scope_key, pid, chip_count)

            chip_total += chip_count
            chip_points = chip_count * int(chip_rate)
            grand_total = score_points + chip_points

            with c5:
                st.write(f"{chip_points:+d}")
            with c6:
                st.write(f"**{grand_total:+d}**")

            final_rows.append({
                "名前": name,
                "累計": total_score,
                f"点数pt（1点={score_rate}pt）": score_points,
                "チップ枚数": chip_count,
                f"チップpt（1枚={chip_rate}pt）": chip_points,
                "総合計": grand_total,
            })

        score_total = sum(int(r["累計"]) for r in final_rows)
        score_point_total = sum(int(r[f"点数pt（1点={score_rate}pt）"]) for r in final_rows)
        chip_point_total = sum(int(r[f"チップpt（1枚={chip_rate}pt）"]) for r in final_rows)
        grand_total_sum = sum(int(r["総合計"]) for r in final_rows)

        # チップ合計のズレは常に見えるように軽く表示
        if chip_total != 0:
            st.error(f"チップ枚数の合計が {chip_total:+d} です。0になるように入力してください。")

        with st.expander("合計チェック", expanded=False):
            st.markdown("---")
            check_col1, check_col2, check_col3 = st.columns(3, gap="small")
            check_col1.metric("累計合計", f"{score_total:+d}")
            check_col2.metric("チップ合計", f"{chip_total:+d}")
            check_col3.metric("総合計", f"{grand_total_sum:+d}")

            sub_col1, sub_col2 = st.columns(2, gap="small")
            sub_col1.metric("点数pt合計", f"{score_point_total:+d}")
            sub_col2.metric("チップpt合計", f"{chip_point_total:+d}")

            if score_total != 0:
                st.error(f"累計合計が {score_total:+d} です。0になっていません。")
            if chip_total != 0:
                st.error(f"チップ枚数の合計が {chip_total:+d} です。0になるように入力してください。")
            if grand_total_sum == 0 and score_total == 0 and chip_total == 0:
                st.success("総合計は0です。")
            elif grand_total_sum != 0:
                st.error(f"総合計が {grand_total_sum:+d} です。0になっていません。")

        st.subheader("総合計ランキング")
        final_rows.sort(key=lambda x: x["総合計"], reverse=True)
        for i, r in enumerate(final_rows, start=1):
            r["順位"] = i
        display_rows = []
        for r in final_rows:
            display_rows.append({
                "順位": r["順位"],
                "名前": r["名前"],
                "累計": r["累計"],
                "点数pt": r[f"点数pt（1点={score_rate}pt）"],
                "チップ枚数": r["チップ枚数"],
                "チップpt": r[f"チップpt（1枚={chip_rate}pt）"],
                "総合計": r["総合計"],
            })
        st.table(display_rows)


# =========================
# 過去の対戦履歴
# =========================
elif st.session_state.page == "history":
    st.title("🕘 過去の対戦履歴")
    back_button()

    sessions = get_sessions(include_finished=True)
    if sessions:
        st.subheader("対戦会一覧")
        rows = []
        for s in sessions:
            sid = s["id"]
            session_results = get_results(session_id=sid)
            ranking = make_enhanced_ranking(session_results)
            top_text = f"{ranking[0]['名前']} {ranking[0]['合計点']:+d}" if ranking else "-"
            rows.append({
                "日付": s.get("session_date"),
                "対戦名": s.get("title"),
                "状態": "進行中" if s.get("status") == "active" else "終了",
                "対戦数": get_session_game_count(sid),
                "1位": top_text,
            })
        st.table(rows)

    st.markdown("---")
    selected_session_id = select_result_scope()
    results = get_results(session_id=selected_session_id)
    history = make_history_table(results)
    if not history:
        st.info("まだ履歴がありません。")
    else:
        st.table(history)


# =========================
# 相性分析
# =========================
elif st.session_state.page == "matchup":
    st.title("🤝 相性分析")
    back_button()
    selected_session_id = select_result_scope()
    results = get_results(session_id=selected_session_id)
    names = sorted(set(r["name"] for r in results if r["name"]))
    if not names:
        st.info("まだデータがありません。")
    else:
        target = st.selectbox("自分を選択", names)
        table = make_matchup_table(results, target)
        if table:
            best = table[0]
            worst = table[-1]
            c1, c2 = st.columns(2)
            c1.metric("相性が良い相手", best["相手"], f"差分 {best['差分']:+d}")
            c2.metric("苦手な相手", worst["相手"], f"差分 {worst['差分']:+d}")
            st.markdown("---")
            st.table(table)
        else:
            st.info("この人の同卓データがまだありません。")


# =========================
# 設定
# =========================
elif st.session_state.page == "settings":
    st.title("⚙️ 設定")
    back_button()

    st.subheader("Excel出力")
    results_for_excel = get_results()
    if results_for_excel:
        excel_bytes = make_excel_file(results_for_excel)
        st.download_button(
            "Excelファイルをダウンロード",
            data=excel_bytes,
            file_name="mahjong_score_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.info("Excel出力できるデータがまだありません。")

    st.markdown("---")
    st.subheader("データ管理")
    st.warning("名前登録データは残したまま、点数一覧・ランキング・個人成績・過去の対戦履歴だけを削除します。")

    game_count = get_game_count()
    result_count = get_result_count()
    st.info(f"現在の対戦数：{game_count}戦 / 点数データ：{result_count}件")

    if st.session_state.clear_scores_step == 0:
        if st.button("点数データを全削除", type="primary", use_container_width=True):
            st.session_state.clear_scores_step = 1
            st.rerun()

    elif st.session_state.clear_scores_step == 1:
        st.subheader("管理者確認")
        admin_pw = st.text_input("管理者パスワード", type="password", key="admin_pw_clear_scores")
        c1, c2 = st.columns(2, gap="small")
        with c1:
            if st.button("確認する", use_container_width=True):
                if admin_pw == ADMIN_PASSWORD:
                    st.session_state.clear_scores_step = 2
                    st.rerun()
                else:
                    st.error("管理者パスワードが違います。")
        with c2:
            if st.button("キャンセル", use_container_width=True):
                st.session_state.clear_scores_step = 0
                st.rerun()

    elif st.session_state.clear_scores_step == 2:
        st.error("本当に点数データを全て削除しますか？この操作は元に戻せません。")
        st.write(f"削除対象：**{game_count}戦** / **{result_count}件**")
        st.write("削除されないもの：**名前登録データ**")

        yes_col, no_col = st.columns(2, gap="small")
        with yes_col:
            if st.button("はい、削除します", use_container_width=True):
                ok, msg = clear_score_data()
                st.session_state.clear_scores_step = 0
                if ok:
                    for key in list(st.session_state.keys()):
                        if key.startswith("manual_point_") or key in ["selected_player_ids", "save_complete", "current_session_id", "setup_session_player_ids"]:
                            del st.session_state[key]
                    st.success(msg)
                    st.rerun()
                else:
                    st.warning(msg)
        with no_col:
            if st.button("いいえ、やめます", use_container_width=True):
                st.session_state.clear_scores_step = 0
                st.rerun()
